import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

DARK_BG      = "1A1A2E"
ORANGE       = "F97316"
GREEN        = "22C55E"
LIGHT_GREEN  = "D1FAE5"
LIGHT_ORANGE = "FFF7ED"
LIGHT_BLUE   = "EFF6FF"
LIGHT_GRAY   = "F8FAFC"
WHITE        = "FFFFFF"
LIGHT_RED    = "FEF2F2"
LIGHT_YELLOW = "FFFBEB"

def s(ws, row, col, value, bold=False, bg=None, fc="1E293B", size=11, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fc, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="E2E8F0")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

# ══ FOGLIO 1 - CALENDARIO ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "Calendario Campagna"
ws1.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [14, 22, 30, 42, 24, 16]):
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 48

ws1.merge_cells("A1:F1")
s(ws1, 1, 1, "CAMPAGNA EMAIL RISTOAGENT - Aprile/Maggio 2026", bold=True, bg=DARK_BG, fc=WHITE, size=16, align="center")

for col, h in enumerate(["DATA", "GIORNO", "AZIONE", "COMANDO DA ESEGUIRE", "TARGET", "STATO"], 1):
    s(ws1, 2, col, h, bold=True, bg=ORANGE, fc=WHITE, align="center")
ws1.row_dimensions[2].height = 28

rows = [
    ("15 Apr", "Oggi",      "TEST email",                  "cd marketing\npython invia_email.py --test",                        "ristoagent@gmail.com",            "Da fare"),
    ("15 Apr", "Oggi",      "Verifica email test",          "Apri Gmail e controlla layout, link e prezzi",                      "ristoagent@gmail.com",            "Da fare"),
    ("16 Apr", "Domani",    "EMAIL 1 - Tier A",             "python invia_email.py --send --step 1 --tier A",                    "64 ristoranti (rating >= 4.5)",   "Da fare"),
    ("22 Apr", "Martedi",   "EMAIL 1 - Tier B",             "python invia_email.py --send --step 1 --tier B",                    "71 ristoranti (rating 4.0-4.4)",  "Da fare"),
    ("22 Apr", "Martedi",   "EMAIL 2 - Follow-up Tier A",   "python invia_email.py --send --step 2 --tier A",                    "Non-aperture Tier A (6gg dopo)",  "Da fare"),
    ("24 Apr", "Giovedi",   "EMAIL 2 - Follow-up Tier B",   "python invia_email.py --send --step 2 --tier B",                    "Non-aperture Tier B (2gg dopo)",  "Da fare"),
    ("29 Apr", "Martedi",   "EMAIL 3 - Last Call",          "python invia_email.py --send --step 3",                             "Tutti i non-convertiti",          "Da fare"),
    ("30 Apr", "Mercoledi", "Analisi risultati",            "resend.com/emails - controlla open rate e click rate",              "Dashboard Resend",                "Da fare"),
]

bgs = [LIGHT_GREEN, LIGHT_GREEN, LIGHT_ORANGE, LIGHT_ORANGE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_RED, LIGHT_GRAY]
for i, (row_data, bg) in enumerate(zip(rows, bgs), 3):
    ws1.row_dimensions[i].height = 44
    data, giorno, azione, cmd, target, stato = row_data
    s(ws1, i, 1, data,    bold=True, bg=bg, align="center")
    s(ws1, i, 2, giorno,  bg=bg, align="center")
    s(ws1, i, 3, azione,  bold=True, bg=bg)
    s(ws1, i, 4, cmd,     bg="1E293B", fc="4ADE80", size=10, wrap=True)
    s(ws1, i, 5, target,  bg=bg, size=10)
    s(ws1, i, 6, "[ ] " + stato, bg=bg, align="center", bold=True, fc=ORANGE)

ws1.row_dimensions[12].height = 24
ws1.merge_cells("A12:F12")
s(ws1, 12, 1, "Orario consigliato: martedi/giovedi ore 9:30-10:30 oppure 14:30-16:00  |  PC acceso solo durante invio (~5 min per sessione)",
  bg=LIGHT_YELLOW, fc="92400E", size=10, align="center")

# ══ FOGLIO 2 - CHECKLIST ════════════════════════════════════════
ws2 = wb.create_sheet("Checklist Pre-Invio")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCD", [6, 35, 52, 16]):
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 45

ws2.merge_cells("A1:D1")
s(ws2, 1, 1, "CHECKLIST - Verifica prima di ogni sessione di invio", bold=True, bg=DARK_BG, fc=WHITE, size=14, align="center")

for col, h in enumerate(["", "COSA VERIFICARE", "DOVE / COME", "STATO"], 1):
    s(ws2, 2, col, h, bold=True, bg=ORANGE, fc=WHITE, align="center")
ws2.row_dimensions[2].height = 26

checks = [
    ("[ ]", "Dominio Resend verificato",              "resend.com/domains -> ristoagent.com deve essere VERDE (Verified)",   "FATTO"),
    ("[ ]", "Email di test inviata e verificata",      "Apri ristoagent@gmail.com e controlla layout + link funzionanti",     "Da fare"),
    ("[ ]", "File CSV presente",                       "marketing/contatti_email.csv - deve esistere (140 righe)",            "FATTO"),
    ("[ ]", "Log invii aggiornato",                    "marketing/invii_log.json - creato automaticamente dallo script",      "Auto"),
    ("[ ]", "Resend API key attiva",                   "resend.com/api-keys - chiave re_dTaGokuT... non scaduta",             "FATTO"),
    ("[ ]", "Pagina /upgrade con Founding Members",    "ristoagent.com/upgrade - sezione arancione visibile in cima",         "FATTO"),
    ("[ ]", "Prezzi Founding Members su Stripe",       "Stripe Dashboard -> Products -> Starter 19 EUR, Pro 29 EUR",          "FATTO"),
    ("[ ]", "PC connesso a internet",                  "Necessario solo durante esecuzione script (~5 minuti)",               "—"),
]

for i, (chk, cosa, dove, stato) in enumerate(checks, 3):
    ws2.row_dimensions[i].height = 38
    bg = LIGHT_GREEN if stato == "FATTO" else LIGHT_ORANGE if stato == "Da fare" else LIGHT_GRAY
    s(ws2, i, 1, chk,  bg=bg, align="center", size=14)
    s(ws2, i, 2, cosa, bold=True, bg=bg)
    s(ws2, i, 3, dove, bg=bg, size=10, wrap=True, fc="374151")
    fc = GREEN if stato == "FATTO" else ORANGE if stato == "Da fare" else "94A3B8"
    s(ws2, i, 4, stato, bg=bg, align="center", bold=True, fc=fc)

# ══ FOGLIO 3 - COMANDI RAPIDI ═══════════════════════════════════
ws3 = wb.create_sheet("Comandi Rapidi")
ws3.sheet_view.showGridLines = False
for col, w in zip("ABC", [28, 60, 28]):
    ws3.column_dimensions[col].width = w
ws3.row_dimensions[1].height = 45

ws3.merge_cells("A1:C1")
s(ws3, 1, 1, "COMANDI RAPIDI - Copia e incolla nel terminale", bold=True, bg=DARK_BG, fc=WHITE, size=14, align="center")

ws3.row_dimensions[2].height = 30
ws3.merge_cells("A2:C2")
s(ws3, 2, 1, "PRIMA DI TUTTO: apri il terminale (cmd) e digita:   cd C:\\Users\\Admin\\OneDrive\\RISTOAGENT\\marketing",
  bg=LIGHT_YELLOW, fc="92400E", size=11, align="center", bold=True)

for col, h in enumerate(["AZIONE", "COMANDO", "NOTE"], 1):
    s(ws3, 3, col, h, bold=True, bg=ORANGE, fc=WHITE, align="center")
ws3.row_dimensions[3].height = 26

cmds = [
    ("Test email (SEMPRE prima)",             "python invia_email.py --test",                                           "Invia 1 email a ristoagent@gmail.com"),
    ("Email 1 - Tier A  (22 apr)",            "python invia_email.py --send --step 1 --tier A",                         "64 ristoranti, rating >= 4.5"),
    ("Email 1 - Tier B  (24 apr)",            "python invia_email.py --send --step 1 --tier B",                         "71 ristoranti, rating 4.0-4.4"),
    ("Email 2 - Follow-up A  (27 apr)",       "python invia_email.py --send --step 2 --tier A",                         "Solo chi non ha aperto Email 1"),
    ("Email 2 - Follow-up B  (29 apr)",       "python invia_email.py --send --step 2 --tier B",                         "Solo chi non ha aperto Email 1"),
    ("Email 3 - Last Call  (5 mag)",          "python invia_email.py --send --step 3",                                  "Tutti i non-convertiti"),
    ("Riprendi se interrotto",                "python invia_email.py --send --step 1 --tier A --start 30",              "Sostituisci 30 con l'indice"),
    ("Quante email inviate?",                 'python -c "import json; d=json.load(open(\'invii_log.json\')); [print(k,len(v)) for k,v in d.items() if isinstance(v,list)]"', "Conta per step"),
]

for i, (azione, cmd, note) in enumerate(cmds, 4):
    ws3.row_dimensions[i].height = 38
    bg = LIGHT_ORANGE if i % 2 == 0 else WHITE
    s(ws3, i, 1, azione, bold=True, bg=bg)
    s(ws3, i, 2, cmd,    bg="1E293B", fc="4ADE80", size=10, wrap=True)
    s(ws3, i, 3, note,   bg=bg, size=10, fc="64748B", wrap=True)

# ══ FOGLIO 4 - METRICHE ════════════════════════════════════════
ws4 = wb.create_sheet("Metriche")
ws4.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [22, 18, 18, 22, 32]):
    ws4.column_dimensions[col].width = w
ws4.row_dimensions[1].height = 45

ws4.merge_cells("A1:E1")
s(ws4, 1, 1, "METRICHE DA MONITORARE - resend.com/emails dopo ogni invio", bold=True, bg=DARK_BG, fc=WHITE, size=14, align="center")

for col, h in enumerate(["METRICA", "TARGET MINIMO", "TARGET BUONO", "IL TUO RISULTATO", "AZIONE SE BASSO"], 1):
    s(ws4, 2, col, h, bold=True, bg=ORANGE, fc=WHITE, align="center")
ws4.row_dimensions[2].height = 26

metrics = [
    ("Open Rate (email aperte)",  "> 20%",   "> 35%",  "— compila dopo invio —", "Cambia oggetto email al prossimo step"),
    ("Click Rate (link cliccati)","> 2%",    "> 5%",   "— compila dopo invio —", "Migliora CTA o aggiungi urgenza"),
    ("Bounce Rate (errori)",      "< 2%",    "< 1%",   "— compila dopo invio —", "Rimuovi email non valide dal CSV"),
    ("Unsubscribe Rate",          "< 0.5%",  "< 0.2%", "— compila dopo invio —", "Riduci frequenza, migliora rilevanza"),
    ("Nuovi iscritti su sito",    "> 2",     "> 5",    "— compila dopo invio —", "Controlla che /upgrade funzioni bene"),
]

for i, (m, tmin, tok, risultato, azione) in enumerate(metrics, 3):
    ws4.row_dimensions[i].height = 38
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    s(ws4, i, 1, m,         bold=True, bg=bg)
    s(ws4, i, 2, tmin,      bg=LIGHT_RED,    align="center", bold=True, fc="991B1B")
    s(ws4, i, 3, tok,       bg=LIGHT_GREEN,  align="center", bold=True, fc="166534")
    s(ws4, i, 4, risultato, bg=LIGHT_YELLOW, align="center", fc="92400E")
    s(ws4, i, 5, azione,    bg=bg, size=10,  fc="64748B", wrap=True)

ws4.row_dimensions[9].height = 28
ws4.merge_cells("A9:E9")
s(ws4, 9, 1, "Dove trovare le metriche: resend.com -> vai su Emails -> clicca su una email inviata -> vedi stats in alto",
  bg=LIGHT_BLUE, fc="1E40AF", size=10, align="center")

path = "C:/Users/Admin/OneDrive/RISTOAGENT/marketing/PROMEMORIA_CAMPAGNA_v2.xlsx"
wb.save(path)
print("File creato:", path)
